import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('Book1.xlsx')
sheet = wb['Sheet1']

cell = sheet['a1'].value
total_row = sheet.max_row
sheet['d1'].value = "Discount price"

for row in range(2, total_row + 1):
    cell = sheet.cell(row, 3).value
    discount_price = cell * 0.9
    discount_price_cell = sheet.cell(row, 4)
    discount_price_cell.value = discount_price
    print(discount_price)

discount_values = Reference(
    sheet,
    min_row=2,
    max_row=sheet.max_row,
    min_col=4,
    max_col=4
)
chart = BarChart()
chart.add_data(discount_values)
sheet.add_chart(chart, 'c6')

wb.save('Book2.xlsx')
